from tkinter import ttk, Toplevel, Label, Button, messagebox
import tkinter as tk
from tkcalendar import Calendar
from decimal import Decimal
from database import get_total_income, get_total_expenses, get_cabin_statistics, get_cabin_statistics_by_date_range, get_total_income_by_date_range, get_total_expenses_by_date_range, fetch_statistics, get_product_sales_statistics_by_dates, get_product_sales_statistics_by_period
from datetime import datetime, timedelta
import pandas as pd
from tkinter import filedialog
from openpyxl.utils import get_column_letter
from tkcalendar import DateEntry  # Добавить в импорты

def create_statistics_page(root):
    """Создает фрейм для отображения статистики."""
    frame = tk.Frame(root)
    current_date_range = {"start": None, "end": None}
    # Таблица статистики по кабинкам
    cabin_frame = tk.Frame(frame)
    cabin_frame.pack(pady=10, fill=tk.X)
    # Заголовок и кнопка для таблицы кабинок
    cabin_header = tk.Frame(cabin_frame)
    cabin_header.pack(fill=tk.X)

    # Заголовок
    tk.Label(cabin_header, text="Статистика по кабинкам", font=("Arial", 16), anchor="center").pack(side=tk.LEFT, expand=True)
    ttk.Button(cabin_header, text="Скачать", command=lambda: export_table_to_excel(tree, "cabin_statistics.xlsx", current_date_range)).pack(side=tk.RIGHT)
    # Таблица статистики по кабинкам
    columns = ("cabin_name", "rental_count", "total_income", "avg_check", "total_people")
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=10)
    tree.heading("cabin_name", text="Название кабинки")
    tree.heading("rental_count", text="Кол-во аренд")
    tree.heading("total_income", text="Общий доход")
    tree.heading("avg_check", text="Средний чек")
    tree.heading("total_people", text="Обслужено людей")
    tree.pack(padx=10, pady=10)
    # Финансовая и продуктовая статистика
    stats_frame = tk.Frame(frame)
    stats_frame.pack(pady=10, fill=tk.X)
    # Финансовая статистика
    finance_frame = tk.Frame(stats_frame)
    finance_frame.pack(side=tk.LEFT, pady=10)

    # Заголовок и кнопка для финансовой статистики
    finance_header = tk.Frame(finance_frame)
    finance_header.pack(fill=tk.X)

    
    tk.Label(finance_header, text="Финансовая статистика", font=("Arial", 14), anchor="center").pack(side=tk.LEFT, expand=True)
    ttk.Button(finance_header, text="Скачать", command=lambda: export_table_to_excel(finance_tree, "finance_statistics.xlsx", current_date_range)).pack(side=tk.RIGHT)

    finance_tree = ttk.Treeview(finance_frame, columns=("description", "value"), show="headings", height=5)
    finance_tree.heading("description", text="Описание")
    finance_tree.heading("value", text="Сумма (₸)")
    finance_tree.pack(padx=10, pady=10)

     # Статистика по продуктам
    product_frame = tk.Frame(stats_frame)
    product_frame.pack(side=tk.RIGHT, padx=10)
    # Заголовок и кнопка для статистики по продуктам
    product_header = tk.Frame(product_frame)
    product_header.pack(fill=tk.X)

    tk.Label(product_header, text="Статистика по продуктам", font=("Arial", 14), anchor="center").pack(side=tk.LEFT, expand=True)
    ttk.Button(product_header, text="Скачать", command=lambda: export_table_to_excel(product_tree, "product_statistics.xlsx", current_date_range)).pack(side=tk.RIGHT)
    product_tree = ttk.Treeview(product_frame, columns=("product_name", "total_sold", "total_income"), show="headings", height=10)
    product_tree.heading("product_name", text="Название продукта")
    product_tree.heading("total_sold", text="Продано (шт.)")
    product_tree.heading("total_income", text="Доход (₸)")
    product_tree.pack(padx=10, pady=10)

    # Кнопки для выбора периода
    button_frame = tk.Frame(frame)
    button_frame.pack(pady=10)

    def update_cabin_statistics(period, date=None):
        """Обновляет данные по кабинкам."""
        if date:
            data = get_cabin_statistics_by_date_range(date[0], date[1])
        else:
            data = get_cabin_statistics(period)

        # Очистка текущих данных
        for item in tree.get_children():
            tree.delete(item)

        total_rentals = 0
        total_income = 0.0
        total_avg_check = 0.0
        total_people_served = 0
        total_cabins = len(data)

        # Заполнение таблицы данными
        for row in data:
            try:
                cabin_name = row[1]  # Название кабинки
                rentals_count = int(row[2]) if row[2] is not None else 0  # Количество аренд
                total_income_row = float(row[3]) if row[3] is not None else 0.0  # Доход
                total_people = int(row[4]) if row[4] is not None else 0
                avg_check = float(row[5]) if row[5] is not None else 0.0  # Средний чек

                tree.insert("", "end", values=(cabin_name, rentals_count, f"{total_income_row:.2f}", f"{avg_check:.2f}", total_people))
                total_rentals += rentals_count
                total_income += total_income_row
                total_people_served += total_people
                total_avg_check += avg_check

            except Exception as e:
                print(f"Ошибка при обработке строки {row}: {e}")

        # Рассчитать общий средний чек
        overall_avg_check = total_income / total_rentals if total_rentals > 0 else 0

        # Добавить пустую строку для отступа
        tree.insert("", "end", values=("", "", "", ""), tags=("spacer",))

        # Добавить строку с итогами
        tree.insert("", "end", values=("Итого", total_rentals, f"{total_income:.2f}", f"{overall_avg_check:.2f}", total_people_served), tags=("summary",))

        # Стилизация итоговой строки и отступа
        tree.tag_configure("summary", font=("Helvetica", 10, "bold"))
        tree.tag_configure("spacer")

    def update_financial_statistics(period, date=None):
        """Обновляет данные финансовой статистики."""
        # Очистка текущих данных
        for item in finance_tree.get_children():
            finance_tree.delete(item)

        if date:
            # Передача двух параметров: start_date и end_date
            start_date, end_date = date
            total_income = get_total_income_by_date_range(start_date, end_date)
            total_expenses = get_total_expenses_by_date_range(start_date, end_date)
        else:
            total_income = get_total_income(period)
            total_expenses = get_total_expenses(period)

        net_profit = Decimal(total_income) - Decimal(total_expenses)

        # Добавляем строки с данными
        finance_tree.insert("", "end", values=("Общий доход", f"{total_income:.2f}"))
        finance_tree.insert("", "end", values=("Общие расходы", f"{total_expenses:.2f}"))

        # Добавляем строку для "Чистой прибыли" с динамическим цветом
        profit_item = finance_tree.insert("", "end", values=("Чистая прибыль", f"{net_profit:.2f}"))

        # Настраиваем стиль для итоговой строки
        finance_tree.tag_configure("summary_negative", foreground="red", font=("Helvetica", 10, "bold"))
        finance_tree.tag_configure("summary_positive", foreground="green", font=("Helvetica", 10, "bold"))
        finance_tree.tag_configure("summary_neutral", foreground="black", font=("Helvetica", 10, "bold"))

        if net_profit < 0:
            finance_tree.item(profit_item, tags="summary_negative")
        elif net_profit > 0:
            finance_tree.item(profit_item, tags="summary_positive")
        else:
            finance_tree.item(profit_item, tags="summary_neutral")

    def update_product_statistics(period, date=None):
        """Обновляет данные статистики по проданным продуктам."""
        # Очистка текущих данных
        for item in product_tree.get_children():
            product_tree.delete(item)

        # Получение данных в зависимости от периода или диапазона дат
        if date:
            start_date, end_date = date
            product_data = get_product_sales_statistics_by_dates(start_date, end_date)
        else:
            product_data = get_product_sales_statistics_by_period(period)

        # Инициализация переменных для подсчета итогов
        total_products_sold = 0
        total_income = 0

        # Добавление данных в виджет Treeview
        for row in product_data:
            product_name, total_sold, income = row
            product_tree.insert("", "end", values=(product_name, total_sold, f"{income:.2f}"))
            
            # Суммирование общего количества проданных товаров и дохода
            total_products_sold += total_sold
            total_income += income

        # Добавить пустую строку для отступа
        product_tree.insert("", "end", values=("", "", ""), tags=("spacer",))

        # Добавить строку с итогами
        product_tree.insert("", "end", values=("Итого", total_products_sold, f"{total_income:.2f}"), tags=("summary",))

        # Стилизация итоговой строки и отступа
        product_tree.tag_configure("summary", font=("Helvetica", 10, "bold"))
        product_tree.tag_configure("spacer")
    
    def open_date_picker(): 
        def select_dates():
            start_date = datetime.strptime(cal_start.get_date(), "%m/%d/%y").strftime("%Y-%m-%d")
            end_date = datetime.strptime(cal_end.get_date(), "%m/%d/%y").strftime("%Y-%m-%d")
            current_date_range["start"] = start_date  # Обновляем start
            current_date_range["end"] = end_date  # Обновляем end
            print(f"Выбранный диапазон: с {start_date} по {end_date}")
            update_statistics(None, (start_date, end_date))
            date_picker.destroy()
            
        date_picker = Toplevel()
        date_picker.title("Выбор диапазона дат")
        date_picker.configure(bg="#e0f7fa")
        
        Label(date_picker, bg="#e6f7ff", fg="black", text="Дата с:").pack()
        cal_start = Calendar(date_picker)
        cal_start.pack()

        Label(date_picker, bg="#e6f7ff", fg="black", text="Дата по:").pack()
        cal_end = Calendar(date_picker)
        cal_end.pack()

        ttk.Button(date_picker, text="Применить", command=select_dates).pack()

    def update_statistics(period, date=None):
        if date:
            start_date, end_date = date
            selected_period_label.config(text=f"Выбранный период: {start_date} - {end_date}")
        else:
            period_text = {"day": "Сегодня", "week": "Эта неделя", "month": "Этот месяц"}
            selected_period_label.config(text=f"Выбранный период: {period_text.get(period, 'Неизвестно')}")

        update_cabin_statistics(period, date)
        update_financial_statistics(period, date)
        update_product_statistics(period, date)

    
    def export_table_to_excel(treeview, default_filename, date_range):
        """Экспортирует данные из Treeview в Excel с указанием даты и позволяет выбрать место сохранения."""
        data = []
        for item in treeview.get_children():
            values = treeview.item(item, 'values')
            data.append(values)

        if not data:
            messagebox.showwarning("Ошибка", "Нет данных для экспорта.")
            return

        # Получаем заголовки столбцов
        columns = [treeview.heading(col)["text"] for col in treeview["columns"]]
        df = pd.DataFrame(data, columns=columns)

        # Окно выбора пути сохранения файла
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
            initialfile=default_filename,
            title="Сохранить файл"
        )

        if not file_path:  # Если пользователь отменил сохранение
            return

        # Создаем Excel-файл
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Статистика")

            # Добавляем строку с датой
            workbook = writer.book
            worksheet = writer.sheets["Статистика"]
            if date_range["start"] and date_range["end"]:
                worksheet.cell(row=1, column=1, value=f"Дата: с {date_range['start']} по {date_range['end']}")
            else:
                worksheet.cell(row=1, column=1, value="Дата: не указана")

            # === АВТОМАТИЧЕСКАЯ ПОДСТРОЙКА ШИРИНЫ КОЛОНОК ===
            for col_num, column_cells in enumerate(worksheet.iter_cols(), start=1):
                max_length = 0
                col_letter = get_column_letter(col_num)  # Получаем букву столбца
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2  # Добавляем немного пространства
                worksheet.column_dimensions[col_letter].width = adjusted_width
                
        messagebox.showinfo("Успех ✅", f"Данные успешно экспортированы в {file_path}.")

    # Кнопки переключения периода
    ttk.Button(button_frame, text="День", command=lambda: update_statistics('day')).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="Неделя", command=lambda: update_statistics('week')).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="Месяц", command=lambda: update_statistics('month')).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="Выбрать дату", command=open_date_picker).pack(side=tk.LEFT, padx=5)
    selected_period_label = tk.Label(button_frame, text="Выбранный период: Сегодня", font=("Arial", 12))
    selected_period_label.pack(side=tk.LEFT, padx=10)

    # Загружаем статистику за день по умолчанию
    update_statistics('day')

    return frame
